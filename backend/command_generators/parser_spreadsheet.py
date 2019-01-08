import io
import mimetypes
import urllib
from typing import List, Optional

import openpyxl
# import koala # An Excel files parser elaborating a graph allowing automatic evaluation

import backend
from backend import Issue
from backend.command_executors import create_command
from backend.command_generators.parser_spreadsheet_utils import binary_mask_from_worksheet, \
    obtain_rectangular_submatrices
from backend.common.helper import create_dictionary, first
from command_definitions import valid_v2_command_names, commands
from command_generators.spreadsheet_command_parsers_v2 import parse_command


def handle_import_commands(r):

    def load_file(location: str = None):
        """
        Loads a case study file (well, really any file) into a BytesIO object
        :param location: URL of the case study file
        :return: bytes
        """
        f_type = None
        data = None

        if location:
            # Try to load the Dataset from the specified location
            data = urllib.request.urlopen(location).read()
            # data = io.BytesIO(data)
            # Then, try to read it
            t = mimetypes.guess_type(location, strict=True)
            if t[0] == "text/python":
                f_type = "python"
            elif t[0] == "text/json":
                f_type = "json"
            elif t[0] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                f_type = "spreadsheet"

        return f_type, data

    workbook = r.get("workbook_name", None)
    worksheets = r.get("worksheets", None)
    sublist2 = [w.strip() for w in worksheets.split(",")] if worksheets else None  # Convert to list of worksheets
    # Read file in memory
    generator_type, file2 = load_file(workbook)
    return generator_type, file2, sublist2


# ############################### #
#  Main function                  #
# ############################### #


def commands_generator_from_ooxml_file(input, state, sublist, stack) -> backend.CommandIssuesPairType:
    """
    It reads an Office Open XML input
    Yields a sequence of command_executors

    :param input: A bytes input
    :param state: State used to check variables
    :param sublist: List of worksheets to consider
    :param stack: Stack of nested files. Just pass it...
    :return:
    """
    # Start the Excel reader
    workbook = openpyxl.load_workbook(io.BytesIO(input), data_only=True)

    # Command names (for the "list of commands" command)
    command_names = create_dictionary(data={cmd_name: None for cmd_name in valid_v2_command_names})

    worksheet_to_command = create_dictionary()  # A dictionary to translate a worksheet to an equivalent command
    if sublist:
        # Force reading "ListOfCommands" commands
        for sheet_name in workbook.sheetnames:
            if commands["ListOfCommands"].regex.search(sheet_name):
                sublist.append(sheet_name)

    # For each worksheet, get the command type, convert into primitive JSON
    for sheet_number, sheet_name in enumerate(workbook.sheetnames):
        if sublist:
            if sheet_name not in sublist:
                continue

        issues = []
        total_issues = []  # type: List[Issue]
        sheet = workbook[sheet_name]

        c_type = None
        c_label = None
        c_content = None

        name = sheet.title

        # Use an equivalent command name
        if name in worksheet_to_command:
            name = worksheet_to_command[name]

        # Extract worksheet matrices
        m = binary_mask_from_worksheet(sheet, False)
        t = obtain_rectangular_submatrices(m, only_remove_empty_bottom=True)
        if len(t) == 0:  # No data
            continue

        t = t[0]  # Take just the first element, a tuple (top, bottom, left, right) representing a rectangular region
        t = (t[0] + 1, t[1] + 1, t[2] + 1, t[3] + 1)  # Indices start at 1

        # v = worksheet_to_numpy_array(sheet)

        # Find which COMMAND to parse, then parse it
        cmd: Optional[backend.Command] = first(commands, condition=lambda c: c.regex.search(name))

        c_type = cmd.name if cmd else None
        if not c_type:
            # Unsupported command
            print(f'Sheet name "{name}" is not a supported command')
            pass

        elif c_type == "etl_dataset":
            if sheet.cell(row=t[0], column=t[2]).value:
                t = (1, m.shape[0] + 1, 1, m.shape[1] + 1)
                # Parse to read parameters
                group2_name = cmd.regex.search(name).group(2)  # Get group name if any
                issues, c_label, c_content = cmd.parse_function(sheet, t, group2_name, state)
            else:
                # Syntax error: it seems there are no parameters
                issue = {"sheet_number": sheet_number, "sheet_name": sheet_name, "c_type": c_type, "type": 3,
                         "message": "It seems there are no parameters for the dataset import command at worksheet '" + sheet_name + "'"}
                total_issues.append(issue)

        elif c_type == "list_of_commands":
            issues, c_label, c_content = cmd.parse_function(sheet, t)
            c_type = None
            if 3 not in [issue.itype for issue in issues]:
                for r in c_content["items"]:
                    worksheet = r.get("worksheet", None)
                    command = r.get("command", None)
                    # Check if valid command
                    if command not in command_names:
                        issue = Issue(sheet_number, sheet_name, 3, None,
                                      "Command '" + command + "' not recognized in List of Commands.")
                    else:
                        worksheet_to_command[worksheet] = command

        elif c_type == "import_commands":
            issues, c_label, c_content = cmd.parse_function(sheet, t)
            if 3 not in [issue.itype for issue in issues]:
                # Declared at this point to avoid circular reference ("parsers_factory" imports "parsers_spreadsheet")
                from backend.command_generators.parsers_factory import commands_container_parser_factory
                # For each line, repeat the import
                for r in c_content["items"]:
                    generator_type, file2, sublist2 = handle_import_commands(r)
                    yield from commands_container_parser_factory(generator_type, None, file2, state, sublist=sublist2, stack=stack)
                    print("Done")

        elif c_type == "mapping":
            groups = cmd.regex.search(name).groups()
            if groups[2] and groups[8]:
                origin = groups[2]
                destination = groups[8]
            elif not groups[2] and not groups[8]:
                origin = None
                destination = None
            else:
                issue = {"sheet_number": sheet_number, "sheet_name": sheet_name, "c_type": c_type, "type": 3,
                         "message": "Either origin or destination are not correctly specified in the sheet name '" + sheet_name + "'"}
                total_issues.append(issue)
            issues, c_label, c_content = cmd.parse_function(sheet, t, origin, destination)

        elif c_type in ["datasetqry", "datasetdata"]:
            issues, c_label, c_content = cmd.parse_function(sheet, t, sheet_name, state)

        elif c_type == "hierarchy":
            res = cmd.regex.search(name)
            h_type = res.group(2)
            c_label = res.group(3)
            issues, _, c_content = cmd.parse_function(sheet, t, c_label, h_type)

        else:
            # GENERIC command parser
            group2_name = cmd.regex.search(name).group(2)  # Get group name if any
            if cmd.parse_function:
                issues, c_label, c_content = cmd.parse_function(sheet, t, group2_name)
            else:
                issues, c_label, c_content = parse_command(sheet, t, group2_name, cmd.label)

        # -------------------------------------------------------------------------------------------------------------
        # Command parsed, now append "issues"
        errors = 0
        if len(issues) > 0:
            for i in issues:
                if isinstance(i, backend.command_generators.Issue):
                    if i.itype == 3:
                        errors += 1
                    issue = {"sheet_number": sheet_number, "sheet_name": sheet_name, "c_type": c_type, "type": i.itype, "message": i.description}
                else:
                    if i[0] == 3:
                        errors += 1
                    issue = {"sheet_number": sheet_number, "sheet_name": sheet_name, "c_type": c_type, "type": i[0], "message": i[1]}
                total_issues.append(issue)
        if errors == 0:
            try:
                if c_type:
                    cmd, issues = create_command(c_type, c_label, c_content, sheet_name)
                else:
                    cmd = None
                    issues = []
            except:
                cmd = None
                issues = [(3, "Could not create command of type '"+c_type+"'")]
            if issues:
                for i in issues:
                    if isinstance(i, backend.command_generators.Issue):
                        issue = {"sheet_number": sheet_number, "sheet_name": sheet_name, "c_type": c_type, "type": i.itype,
                                 "message": i.description}
                    else:
                        issue = {"sheet_number": sheet_number, "sheet_name": sheet_name, "c_type": c_type, "type": i[0],
                                 "message": i[1]}

                    total_issues.append(issue)
        else:
            print(issues)  # Convenient for debugging purposes
            cmd = None  # cmd, _ = create_command(c_type, c_label, {}, sh_name)

        yield cmd, total_issues
    # yield from []  # Empty generator


# def get_codes_all_statistical_datasets(source, dataset_manager):
#     """
#     Obtain a list of datasets available from a source
#     If no source is specified, all the sources are queried
#     For each dataset, the source, the name, the periods available, an example command and a description are obtained
#
#     :param source:
#     :param dataset_manager: It is a DataSourceManager
#     :return: A Dataframe with the list of datasets
#     """
#     lst2 = []
#     # TODO Probably "get_datasets" will not work as expected. It returns a tuple (Source, list of datasets)
#     for r, k in enumerate(dataset_manager.get_datasets(source)):
#         if len(k) == 4:
#             src = k[3]
#         else:
#             src = ""
#         lst2.append((k[0], k[1], k[2], src))
#     return pd.DataFrame(data=lst2, columns=["Dataset ID", "Description", "URN", "Data Source"])

